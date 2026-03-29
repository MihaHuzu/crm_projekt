# export_import.py

import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Blueprint, send_file, request, redirect, url_for, flash, render_template
from flask_login import login_required
from models import db, Customer

# Blueprint - grup de rute pentru export/import
export_import = Blueprint('export_import', __name__)


# ===============================================================
# CSV EXPORT - CUSTOMERS
# ===============================================================

@export_import.route('/export/customers/csv')
@login_required
def export_customers_csv():
    """Exportiert alle Kunden als CSV-Datei"""

    # StringIO = ein virtuelles Textdokument im Arbeitsspeicher
    output = io.StringIO()
    writer = csv.writer(output)

    # Kopfzeile schreiben
    writer.writerow(['id', 'name', 'email', 'company', 'phone', 'status', 'created_at'])

    # Alle Kunden aus der Datenbank holen und schreiben
    customers = Customer.query.all()
    for c in customers:
        writer.writerow([c.id, c.name, c.email, c.company, c.phone, c.status, c.created_at])

    # Cursor zum Anfang zurücksetzen
    output.seek(0)

    # BytesIO = konvertiert Text zu Bytes
    bytes_output = io.BytesIO()
    bytes_output.write(output.getvalue().encode('utf-8-sig'))
    bytes_output.seek(0)

    return send_file(
        bytes_output,
        mimetype='text/csv',
        as_attachment=True,
        download_name='customers.csv'
    )


# ===============================================================
# EXCEL EXPORT - CUSTOMERS
# ===============================================================

@export_import.route('/export/customers/excel')
@login_required
def export_customers_excel():
    """Exportiert alle Kunden als Excel-Datei"""

    # Neue Excel-Arbeitsmappe erstellen
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Customers'

    # Stil für die Kopfzeile
    header_font  = Font(bold=True, color='FFFFFF')
    header_fill  = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    header_align = Alignment(horizontal='center')

    # Kopfzeile schreiben und formatieren
    headers = ['ID', 'Name', 'Email', 'Company', 'Phone', 'Status', 'Created At']
    for col, header in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    # Daten schreiben
    customers = Customer.query.all()
    for row, c in enumerate(customers, start=2):
        ws.cell(row=row, column=1, value=c.id)
        ws.cell(row=row, column=2, value=c.name)
        ws.cell(row=row, column=3, value=c.email)
        ws.cell(row=row, column=4, value=c.company)
        ws.cell(row=row, column=5, value=c.phone)
        ws.cell(row=row, column=6, value=c.status)
        ws.cell(row=row, column=7, value=str(c.created_at))

    # Spaltenbreite automatisch anpassen
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    # In BytesIO speichern
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='customers.xlsx'
    )


# ===============================================================
# HILFSFUNKTIONEN - Validierung
# ===============================================================

def validate_customer_row(row, row_number):
    """Überprüft ob eine Zeile alle erforderlichen Felder enthält"""
    required_fields = ['name', 'email', 'company', 'phone']

    for field in required_fields:
        if field not in row or not row[field].strip():
            return f'Zeile {row_number}: Feld "{field}" fehlt oder ist leer'

    if '@' not in row['email']:
        return f'Zeile {row_number}: Ungültige E-Mail "{row["email"]}"'

    return None  # kein Fehler


# ===============================================================
# DATEI LESEN
# ===============================================================

def read_csv(file):
    """Liest eine CSV-Datei und gibt eine Liste von Dictionaries zurück"""
    try:
        stream = io.StringIO(file.stream.read().decode('utf-8-sig'))
        reader = csv.DictReader(stream)  # jede Zeile als Dictionary
        rows = [row for row in reader]
        return rows, None
    except Exception as e:
        return [], str(e)


def read_excel(file):
    """Liest eine Excel-Datei und gibt eine Liste von Dictionaries zurück"""
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        rows    = []
        headers = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                # Erste Zeile = Kopfzeile
                headers = [str(cell).lower().strip() if cell else '' for cell in row]
            else:
                # Datenzeilen als Dictionary
                row_dict = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = str(value) if value is not None else ''
                rows.append(row_dict)

        return rows, None
    except Exception as e:
        return [], str(e)


def import_customers(rows):
    """Importiert Kunden aus einer Liste von Dictionaries"""
    success = 0
    errors  = []

    for i, row in enumerate(rows, start=2):  # start=2 weil Zeile 1 die Kopfzeile ist
        # Validierung
        error = validate_customer_row(row, i)
        if error:
            errors.append(error)
            continue  # nächste Zeile überspringen

        # Prüfen ob E-Mail bereits existiert
        if Customer.query.filter_by(email=row['email'].strip()).first():
            errors.append(f'Zeile {i}: E-Mail "{row["email"]}" existiert bereits - übersprungen')
            continue

        # Kunden erstellen
        new_customer = Customer(
            name=row['name'].strip(),
            email=row['email'].strip(),
            company=row['company'].strip(),
            phone=row['phone'].strip(),
            status=row.get('status', 'prospect').strip() or 'prospect'
        )
        db.session.add(new_customer)
        success += 1

    db.session.commit()
    return success, errors


# ===============================================================
# IMPORT ROUTE
# ===============================================================

@export_import.route('/import', methods=['GET', 'POST'])
@login_required
def import_data():
    """Import-Seite für CSV und Excel Dateien"""

    if request.method == 'POST':
        file = request.files.get('file')

        # Validierung
        if not file or file.filename == '':
            flash('Keine Datei ausgewählt!', 'error')
            return redirect(url_for('export_import.import_data'))

        filename = file.filename.lower()

        # Dateiformat bestimmen und lesen
        if filename.endswith('.csv'):
            rows, error = read_csv(file)
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            rows, error = read_excel(file)
        else:
            flash('Nur CSV und Excel Dateien erlaubt!', 'error')
            return redirect(url_for('export_import.import_data'))

        if error:
            flash(f'Fehler beim Lesen der Datei: {error}', 'error')
            return redirect(url_for('export_import.import_data'))

        # Kunden importieren
        success, errors = import_customers(rows)

        # Ergebnis anzeigen
        if errors:
            for err in errors:
                flash(err, 'error')

        if success > 0:
            flash(f'{success} Kunden erfolgreich importiert!', 'success')

        return redirect(url_for('export_import.import_data'))

    return render_template('import.html')